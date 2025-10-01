"""
Business: Import Excel files with owner reports data into database
Args: event - dict with httpMethod, body (base64 encoded Excel file), headers (X-Admin-Token)
      context - object with attributes: request_id, function_name
Returns: HTTP response with import results
"""

import json
import base64
import os
from typing import Dict, Any, List
from io import BytesIO
import openpyxl
from datetime import datetime
import psycopg2
from psycopg2.extras import execute_values

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'POST')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, X-Admin-Token',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    if method != 'POST':
        return {
            'statusCode': 405,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Method not allowed'}),
            'isBase64Encoded': False
        }
    
    headers = event.get('headers', {})
    admin_token = headers.get('X-Admin-Token') or headers.get('x-admin-token')
    
    if not admin_token:
        return {
            'statusCode': 401,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': 'Admin token required'}),
            'isBase64Encoded': False
        }
    
    try:
        body_data = json.loads(event.get('body', '{}'))
        file_base64 = body_data.get('file')
        apartment_number = body_data.get('apartment_number')
        
        if not file_base64:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': 'File data required'}),
                'isBase64Encoded': False
            }
        
        if not apartment_number:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': 'Apartment number required'}),
                'isBase64Encoded': False
            }
        
        file_bytes = base64.b64decode(file_base64)
        
        workbook = openpyxl.load_workbook(BytesIO(file_bytes))
        sheet = workbook.active
        
        reports_data: List[tuple] = []
        
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row and any(cell for cell in row):
                if 'Заселение' in str(row) or 'заселение' in str(row).lower():
                    header_row = idx
                    break
        
        if not header_row:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': 'Header row not found'}),
                'isBase64Encoded': False
            }
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(cell for cell in row):
                continue
            
            try:
                check_in = row[0]
                check_out = row[1]
                
                if not check_in or not check_out:
                    continue
                
                if isinstance(check_in, str):
                    check_in = datetime.strptime(check_in, '%d.%m.%Y').date()
                elif hasattr(check_in, 'date'):
                    check_in = check_in.date()
                
                if isinstance(check_out, str):
                    check_out = datetime.strptime(check_out, '%d.%m.%Y').date()
                elif hasattr(check_out, 'date'):
                    check_out = check_out.date()
                
                def parse_number(value):
                    if value is None or value == '':
                        return 0
                    if isinstance(value, (int, float)):
                        return float(value)
                    if isinstance(value, str):
                        value = value.replace(' ', '').replace(',', '.')
                        if value.endswith('%'):
                            return float(value[:-1])
                        return float(value) if value else 0
                    return 0
                
                booking_sum = parse_number(row[5] if len(row) > 5 else 0)
                total_sum = parse_number(row[6] if len(row) > 6 else 0)
                commission_percent = parse_number(row[7] if len(row) > 7 else 0)
                usn_percent = parse_number(row[8] if len(row) > 8 else 0)
                commission_before_usn = parse_number(row[9] if len(row) > 9 else 0)
                commission_after_usn = parse_number(row[10] if len(row) > 10 else 0)
                remaining_before_expenses = parse_number(row[11] if len(row) > 11 else 0)
                expenses_on_operations = parse_number(row[12] if len(row) > 12 else 0)
                owner_payment = parse_number(row[13] if len(row) > 13 else 0)
                average_cleaning = parse_number(row[17] if len(row) > 17 else 0)
                hot_water = parse_number(row[14] if len(row) > 14 else 0)
                chemical_cleaning = parse_number(row[15] if len(row) > 15 else 0)
                hygiene = parse_number(row[19] if len(row) > 19 else 0)
                transportation = parse_number(row[20] if len(row) > 20 else 0)
                utilities = parse_number(row[18] if len(row) > 18 else 0)
                other = parse_number(row[22] if len(row) > 22 else 0)
                note = str(row[25]) if len(row) > 25 and row[25] else None
                
                reports_data.append((
                    apartment_number,
                    check_in,
                    check_out,
                    booking_sum,
                    total_sum,
                    commission_percent,
                    usn_percent,
                    commission_before_usn,
                    commission_after_usn,
                    remaining_before_expenses,
                    expenses_on_operations,
                    average_cleaning,
                    owner_payment,
                    hot_water,
                    chemical_cleaning,
                    hygiene,
                    transportation,
                    utilities,
                    other,
                    note
                ))
                
            except Exception as e:
                continue
        
        if not reports_data:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': 'No valid data found in file'}),
                'isBase64Encoded': False
            }
        
        dsn = os.environ.get('DATABASE_URL')
        conn = psycopg2.connect(dsn)
        cur = conn.cursor()
        
        insert_query = """
            INSERT INTO t_p9202093_hotel_design_site.owner_reports 
            (apartment_number, check_in_date, check_out_date, booking_sum, total_sum, 
             commission_percent, usn_percent, commission_before_usn, commission_after_usn,
             remaining_before_expenses, expenses_on_operations, average_cleaning, 
             owner_payment, hot_water, chemical_cleaning, hygiene_ср_ва, 
             transportation, utilities, other, note_to_billing)
            VALUES %s
        """
        
        execute_values(cur, insert_query, reports_data)
        conn.commit()
        
        inserted_count = len(reports_data)
        
        cur.close()
        conn.close()
        
        return {
            'statusCode': 200,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({
                'success': True,
                'message': f'Imported {inserted_count} reports',
                'count': inserted_count
            }),
            'isBase64Encoded': False
        }
        
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': str(e)}),
            'isBase64Encoded': False
        }
